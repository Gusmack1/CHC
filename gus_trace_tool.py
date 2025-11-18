import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk, Listbox, END
from tkinter import simpledialog
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.chart import PieChart, Reference
import time
import threading
import datetime
import re
from collections import deque
import webbrowser
import pickle
import os
import json
import argparse
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from fuzzywuzzy import fuzz
import configparser

CONFIG_FILE = 'config.ini'
SETTINGS_FILE = 'settings.json'

def load_config():
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
        return config
    return None

_config = load_config()

# Optional hard-coded default API key for private use only.
# For any public/shared version of this project, leave this empty and provide
# the key via environment variable or config.ini instead.
HARDCODED_GUS_API_KEY = ''

# Prefer environment variable, then config file, then the hard-coded default.
GUS_API_KEY = os.getenv('GUS_API_KEY', '')
if not GUS_API_KEY and _config and _config.has_option('API', 'GUS_API_KEY'):
    GUS_API_KEY = _config.get('API', 'GUS_API_KEY').strip()
if not GUS_API_KEY:
    GUS_API_KEY = HARDCODED_GUS_API_KEY

BASE_URL = 'https://api.company-information.service.gov.uk'
REQUEST_TIMEOUT = 10
RATE_LIMIT_WINDOW = 300   # 5 minutes
RATE_LIMIT_MAX = 600      # 600 calls per 5 minutes (Companies House guideline)
CACHE_TIMEOUT = 24 * 3600

# -------- Matching / scoring thresholds (tunable) --------

# Minimum overall accuracy for a candidate to be considered.
MIN_OVERALL_ACCURACY = 60.0

# Penalties / bonuses on company profile.
NON_ACTIVE_STATUS_PENALTY = 10.0
OLD_COMPANY_YEARS = 10
OLD_COMPANY_PENALTY = 5.0

# Address similarity.
ADDRESS_LOW_SIM_THRESHOLD = 0.7   # 70%
ADDRESS_LOW_SIM_PENALTY = 20.0

# Postcode weighting.
POSTCODE_MATCH_BONUS = 15.0
POSTCODE_MISMATCH_PENALTY = 10.0

# Distinctive word scoring.
ALL_DISTINCT_PRESENT_BONUS = 5.0
DISTINCT_MISSING_PENALTY_PER_WORD = 8.0
FIRST_DISTINCT_EXTRA_PENALTY = 15.0

# Prefix alignment (matching leading words in order).
PREFIX_ALIGNMENT_BONUS_PER_WORD = 5.0

APP_NAME = "Companies House Lookup"
APP_VERSION = "1.0.0"


class ToolTip:
    """Simple tooltip for Tkinter widgets."""

    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        self.id = None
        self.widget.bind("<Enter>", self._enter)
        self.widget.bind("<Leave>", self._leave)
        self.widget.bind("<ButtonPress>", self._leave)

    def _enter(self, event=None):
        self._schedule()

    def _leave(self, event=None):
        self._unschedule()
        self._hide_tip()

    def _schedule(self):
        self._unschedule()
        self.id = self.widget.after(600, self._show_tip)

    def _unschedule(self):
        if self.id:
            self.widget.after_cancel(self.id)
            self.id = None

    def _show_tip(self, event=None):
        if self.tipwindow or not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 2
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw,
            text=self.text,
            justify=tk.LEFT,
            background="#ffffe0",
            relief=tk.SOLID,
            borderwidth=1,
            font=("tahoma", 9),
            padx=4,
            pady=2,
        )
        label.pack(ipadx=1)

    def _hide_tip(self):
        tw = self.tipwindow
        if tw:
            tw.destroy()
        self.tipwindow = None

def load_settings():
    """Load persisted GUI settings from JSON, if present."""
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_settings(data):
    """Persist GUI settings to JSON."""
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
    except Exception:
        # Failing to save settings is non-fatal; ignore.
        pass

insolvency_types = {
    "creditors-voluntary-liquidation": "Creditors' Voluntary Liquidation",
    "members-voluntary-liquidation": "Members' Voluntary Liquidation",
    "in-administration": "In Administration",
    "administrative-receivership": "Administrative Receivership",
    "voluntary-arrangement": "Voluntary Arrangement",
    "company-voluntary-arrangement": "Company Voluntary Arrangement",
    "receivership": "Receivership",
    "ptso": "Proposal to Strike Off",
    "dissolved": "Dissolved",
    "liquidation": "Liquidation",
    "administration": "Administration",
    "insolvency-proceedings": "Insolvency Proceedings",
    "receivership": "Receivership",
    "voluntary-arrangement": "Voluntary Arrangement"
}

COMPANY_KEYWORDS = [
    "LTD", "LIMITED", "CO", "COMPANY", "ASSOCIATION", "GROUP", "PROPERTIES", "HOMES", "HOME", "DEVELOPMENTS", "DEVELOPMENT",
    "INVESTMENTS", "TRUST", "LLP", "PLC", "INC", "CORP", "ESTATE", "HOUSING", "MANAGEMENT", "REAL",
    "INVESTMENT", "VENTURES", "CONSTRUCTION", "BUILD", "RENTAL", "PUB", "SCHOOL", "CHURCH", "COUNCIL",
    "NHS", "TRUSTEE", "PARTNERSHIP", "ENTERPRISE", "CIC", "LP", "CARE", "RESIDENTS", "RTM", "PROJECT", "PHASE",
    "SERVICES", "SOLUTIONS", "TECHNOLOGIES", "LABS", "STUDIOS", "AGENCY", "CONSULTING", "HOLDINGS", "TRADING", "ENTERPRISES"
]

def is_likely_company(name):
    name_stripped = name.strip()
    name_upper = name_stripped.upper()

    # Obvious non-company patterns first
    # 1) Email addresses
    if "@" in name_upper:
        return False

    # 2) Pure UK postcode (e.g. EX5 1AF). These are not company names.
    postcode_candidate = re.sub(r"\s+", "", name_upper)
    if re.match(r"^[A-Z]{1,2}\d[A-Z\d]?\d[A-Z]{2}$", postcode_candidate):
        return False

    # 3) Plain schools without company markers – treat as non-company
    # e.g. "STANTON VALE SCHOOL" should not be forced to match a random LTD.
    if "SCHOOL" in name_upper:
        has_company_marker = any(
            marker in name_upper
            for marker in [" LTD", " LIMITED", " TRUST", " ACADEMY", " CIC", " PLC", " LLP"]
        )
        if not has_company_marker and not re.search(r'\d', name_upper):
            return False
    titles = ["MR", "MRS", "MS", "MISS", "DR", "PROF", "SIR", "LADY"]
    parts = name_upper.split()
    if name_upper.startswith(tuple(t + " " for t in titles)) and not any(keyword in name_upper for keyword in COMPANY_KEYWORDS) and not re.search(r'\d', name_upper):
        return False
    if re.search(r'\d', name_upper):
        return True
    if any(keyword in name_upper for keyword in COMPANY_KEYWORDS):
        return True
    if len(parts) > 2:
        return True
    return False

def clean_name(name):
    # Strip trading-as fragments: keep the legal entity part only
    name = re.sub(r'\s+TRADING\s+AS.*$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+T/A\s+.*$', '', name, flags=re.IGNORECASE)

    name = re.sub(r'\s*-\s*', '-', name.strip().upper())

    # Handle truncated 'L' at end that should really be 'LTD'
    # e.g. 'UK INTERIOR SOLUTIONS L' -> 'UK INTERIOR SOLUTIONS LTD'
    name = re.sub(r'\bL$', ' LTD', name)
    name = re.sub(r'\s+', ' ', name)
    parts = name.split()
    i = 0
    while i < len(parts) - 1:
        if len(parts[i]) == 1 and len(parts[i+1]) == 1:
            parts[i] += parts.pop(i+1)
        else:
            i += 1
    if len(parts) > 1 and len(parts) % 2 == 0:
        half = len(parts) // 2
        if parts[:half] == parts[half:]:
            parts = parts[:half]
    unique_parts = []
    for p in parts:
        if not unique_parts or p != unique_parts[-1]:
            unique_parts.append(p)
    name = ' '.join(unique_parts)
    return name

def core_name(name):
    name_upper = name.upper()
    for keyword in COMPANY_KEYWORDS:
        if name_upper.endswith(f' {keyword}'):
            return name_upper[:-len(keyword)-1].strip()
    return name_upper

class RateLimiter:
    def __init__(self, max_requests, time_window):
        self.max_requests = max_requests
        self.time_window = time_window
        self.requests = deque()
        self.lock = Lock()
        self.pauses = 0

    def acquire(self, log_func=None):
        """
        Ensure we stay within Companies House rate limits.
        Optionally logs when we have to pause.
        """
        while True:
            with self.lock:
                current_time = time.time()
                while self.requests and self.requests[0] < current_time - self.time_window:
                    self.requests.popleft()
                if len(self.requests) < self.max_requests:
                    self.requests.append(current_time)
                    return
                sleep_time = self.requests[0] + self.time_window - current_time
            sleep_time = max(sleep_time, 0.1)
            if log_func and sleep_time > 0.5:
                # Surface a clear log line when we deliberately pause.
                log_func(
                    f"Rate limiter: sleeping {sleep_time:.1f}s to respect Companies House limits\n",
                    "yellow",
                )
            time.sleep(sleep_time)
            self.pauses += 1

rate_limiter = RateLimiter(RATE_LIMIT_MAX, RATE_LIMIT_WINDOW)

profile_cache = {}
officers_cache = {}
insolvency_cache = {}
filing_cache = {}
disqualified_cache = {}
search_cache = {}

# GUI state variables that are initialised in create_gui()
verbose_var = None

def load_cache(file_name):
    if os.path.exists(file_name):
        mod_time = os.path.getmtime(file_name)
        if time.time() - mod_time < CACHE_TIMEOUT:
            with open(file_name, 'rb') as f:
                return pickle.load(f)
    return {}

def save_caches():
    with open('profile_cache.pickle', 'wb') as f:
        pickle.dump(profile_cache, f)
    with open('officers_cache.pickle', 'wb') as f:
        pickle.dump(officers_cache, f)
    with open('insolvency_cache.pickle', 'wb') as f:
        pickle.dump(insolvency_cache, f)
    with open('filing_cache.pickle', 'wb') as f:
        pickle.dump(filing_cache, f)
    with open('disqualified_cache.pickle', 'wb') as f:
        pickle.dump(disqualified_cache, f)
    with open('search_cache.pickle', 'wb') as f:
        pickle.dump(search_cache, f)

def clear_all_caches(delete_files=True):
    """Clear all in-memory caches and optionally delete any on-disk cache files."""
    global profile_cache, officers_cache, insolvency_cache, filing_cache, disqualified_cache, search_cache
    profile_cache.clear()
    officers_cache.clear()
    insolvency_cache.clear()
    filing_cache.clear()
    disqualified_cache.clear()
    search_cache.clear()
    if delete_files:
        for fname in [
            'profile_cache.pickle',
            'officers_cache.pickle',
            'insolvency_cache.pickle',
            'filing_cache.pickle',
            'disqualified_cache.pickle',
            'search_cache.pickle',
        ]:
            try:
                if os.path.exists(fname):
                    os.remove(fname)
            except Exception:
                pass

api_calls = 0
api_times = []

def api_get(url, api_key, log_func=None, stop_event=None, pause_event=None):
    global api_calls, api_times
    api_calls += 1
    rate_limiter.acquire(log_func)
    if stop_event and stop_event.is_set():
        raise Exception("Processing stopped")
    start_time = time.time()
    retries = 5
    for attempt in range(retries):
        if stop_event and stop_event.is_set():
            raise Exception("Processing stopped")
        try:
            # If a pause has been requested, wait here until it is cleared or stop is requested.
            while pause_event and pause_event.is_set():
                if stop_event and stop_event.is_set():
                    raise Exception("Processing stopped")
                time.sleep(0.1)

            response = requests.get(url, auth=(api_key, ''), timeout=REQUEST_TIMEOUT)
            if response.status_code == 429:
                # Companies House is still throttling us even after our local rate limiter.
                # Respect Retry-After if present, otherwise use a shorter, increasing backoff.
                retry_after_header = response.headers.get('Retry-After')
                if retry_after_header:
                    try:
                        sleep_for = float(retry_after_header)
                    except ValueError:
                        sleep_for = min(RATE_LIMIT_WINDOW, 60 + attempt * 30)
                else:
                    sleep_for = min(RATE_LIMIT_WINDOW, 60 + attempt * 30)
                if log_func:
                    log_func(f"Rate limit hit on {url}, sleeping {sleep_for:.1f}s\n", "red")
                time.sleep(sleep_for)
                continue
            if response.status_code == 200:
                api_times.append(time.time() - start_time)
                return response.json()
            elif 400 <= response.status_code < 500:
                api_times.append(time.time() - start_time)
                return None
            else:
                if log_func:
                    log_func(f"API error {response.status_code} for {url}\n", "red")
        except requests.exceptions.Timeout:
            if log_func:
                log_func(f"Timeout for {url}, attempt {attempt+1}/{retries}\n", "yellow")
            # Simple exponential backoff on timeouts: 2, 4, 8, ... seconds.
            time.sleep(2 ** attempt)
        except Exception as e:
            if log_func:
                log_func(f"API exception for {url}: {str(e)}\n", "red")
            break
    api_times.append(time.time() - start_time)
    return None

def search_companies(query, api_key, limit=10, log_func=None, stop_event=None, use_cache=True, pause_event=None):
    key = query + str(limit)
    if use_cache and key in search_cache:
        return search_cache[key]
    data = api_get(f"{BASE_URL}/search/companies?q={requests.utils.quote(query)}&items_per_page={limit}", api_key, log_func, stop_event, pause_event)
    items = data.get('items', []) if data else []
    if use_cache:
        search_cache[key] = items
    return items

def get_company_profile(company_number, api_key, log_func=None, stop_event=None, use_cache=True, pause_event=None):
    if use_cache and company_number in profile_cache:
        return profile_cache[company_number]
    data = api_get(f"{BASE_URL}/company/{company_number}", api_key, log_func, stop_event, pause_event)
    if use_cache:
        profile_cache[company_number] = data
    return data

def get_officers(company_number, api_key, log_func=None, stop_event=None, use_cache=True, pause_event=None):
    if use_cache and company_number in officers_cache:
        return officers_cache[company_number]
    data = api_get(f"{BASE_URL}/company/{company_number}/officers", api_key, log_func, stop_event, pause_event)
    officers = data.get('items', []) if data else []
    if use_cache:
        officers_cache[company_number] = officers
    return officers

def get_insolvency(company_number, api_key, log_func=None, stop_event=None, use_cache=True, pause_event=None):
    if use_cache and company_number in insolvency_cache:
        return insolvency_cache[company_number]
    data = api_get(f"{BASE_URL}/company/{company_number}/insolvency", api_key, log_func, stop_event, pause_event)
    if data and 'cases' in data and data['cases']:
        case = data['cases'][0]
        ins_type = case.get('type', '')
        ins_date = case.get('dates', [{}])[0].get('date', '')
    else:
        ins_type, ins_date = '', ''
    if use_cache:
        insolvency_cache[company_number] = (ins_type, ins_date)
    return ins_type, ins_date

def get_filing_history(company_number, api_key, log_func=None, stop_event=None, use_cache=True, pause_event=None):
    if use_cache and company_number in filing_cache:
        return filing_cache[company_number]
    data = api_get(f"{BASE_URL}/company/{company_number}/filing-history", api_key, log_func, stop_event, pause_event)
    filings = data.get('items', []) if data else []
    if use_cache:
        filing_cache[company_number] = filings
    return filings

def get_disqualified_info(name, api_key, log_func=None, stop_event=None, use_cache=True, pause_event=None):
    key = name.upper()
    if use_cache and key in disqualified_cache:
        return disqualified_cache[key]
    data = api_get(f"{BASE_URL}/search/disqualified-officers?q={requests.utils.quote(name)}", api_key, log_func, stop_event, pause_event)
    if data:
        items = data.get('items', [])
        for item in items:
            if item.get('kind') == 'disqualified-officer-natural':
                full_name = ' '.join(filter(None, [item.get('title'), item.get('forename'), item.get('other_forenames'), item.get('surname')])).upper()
                if fuzz.ratio(name.upper(), full_name) > 95:
                    disqs = item.get('disqualifications', [])
                    active_disqs = [d for d in disqs if datetime.datetime.strptime(d.get('disqualified_until', '1900-01-01'), '%Y-%m-%d').date() > datetime.date.today()]
                    if active_disqs:
                        details = [f"Reason: {d.get('reason', {}).get('description_identifier', 'Unknown')}, From: {d.get('disqualified_from', 'N/A')}, Until: {d.get('disqualified_until', 'N/A')}" for d in active_disqs]
                        if use_cache:
                            disqualified_cache[key] = (True, '; '.join(details))
                        return True, '; '.join(details)
    if use_cache:
        disqualified_cache[key] = (False, '')
    return False, ''

def calculate_similarity(str1, str2):
    wratio = fuzz.WRatio(str1.upper(), str2.upper()) / 100.0
    token_sort = fuzz.token_sort_ratio(str1.upper(), str2.upper()) / 100.0
    return max(wratio, token_sort)

def prepare_query(debtor_name):
    cleaned_query = clean_name(debtor_name)
    core_query = core_name(cleaned_query)
    if cleaned_query.endswith(' GROUP'):
        search_query = cleaned_query[:-6].strip()
    else:
        search_query = core_query
    return cleaned_query, core_query, search_query


def extract_postcode(text):
    """
    Extract and normalise a UK postcode from free text.
    Returns an uppercase postcode with no internal spaces, or '' if none.
    """
    if not text:
        return ''
    m = re.search(r'([A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2})', str(text).upper())
    if not m:
        return ''
    return re.sub(r'\s+', '', m.group(1))

def deduplicate_letters(s):
    return re.sub(r'(.)\1+', r'\1', s)

def search_and_variants(search_query, cleaned_query, api_key, log_func, stop_event, use_cache, pause_event=None, max_variants=4):
    """
    Build a small, well-ordered set of search variants, always including:
    - the core search_query
    - the cleaned full name
    - simple LTD/LIMITED swaps and deduped versions
    """
    raw_variants = []

    def add_variant(v):
        v = v.strip()
        if v and v not in raw_variants:
            raw_variants.append(v)

    # Always include the core query and the cleaned full name first.
    add_variant(search_query)
    add_variant(cleaned_query)

    dedup = deduplicate_letters(search_query)
    if dedup != search_query:
        add_variant(dedup)

    if cleaned_query.endswith(' LTD'):
        add_variant(cleaned_query[:-4].strip() + ' LIMITED')
    elif cleaned_query.endswith(' LIMITED'):
        add_variant(cleaned_query[:-8].strip() + ' LTD')

    if 'RTM CO' in cleaned_query.upper():
        add_variant(cleaned_query.upper().replace('RTM CO', 'RTM COMPANY'))
    words = search_query.split()
    new_words = []
    i = 0
    while i < len(words):
        new_words.append(words[i])
        if i < len(words) - 1 and re.search(r'\d$', words[i]):
            new_words[-1] += words[i+1]
            i += 1
        i += 1
    variant = ' '.join(new_words)
    if variant != search_query:
        add_variant(variant)

    # Handle common truncation 'CHIN' -> 'CHINA' for pottery/ceramics companies
    # e.g. 'CAVERSWALL ENGLISH CHIN' -> 'CAVERSWALL ENGLISH CHINA'
    if variant.endswith(" CHIN"):
        chin_base = variant[:-5] + " CHINA"
        add_variant(chin_base)
        add_variant(chin_base + " COMPANY LIMITED")

    variant_singular = re.sub(r'S\b', '', variant)
    if variant_singular != search_query:
        add_variant(variant_singular)

    variants = raw_variants[:max_variants]
    if log_func:
        log_func(f"Search variants for '{search_query}': {', '.join(variants)}\n", "yellow")
    all_results_dict = {}
    with ThreadPoolExecutor(max_workers=min(3, len(variants))) as executor:
        futures = {}
        for v in variants:
            future = executor.submit(search_companies, v, api_key, log_func=log_func, stop_event=stop_event, use_cache=use_cache, pause_event=pause_event)
            futures[future] = v
        for future in as_completed(futures):
            if stop_event.is_set():
                raise Exception("Processing stopped")
            results = future.result()
            for res in results:
                cn = res.get('company_number')
                if cn and cn not in all_results_dict:
                    all_results_dict[cn] = res
    all_results = list(all_results_dict.values())
    if log_func:
        log_func(f"Found {len(all_results)} unique Companies House search results for '{search_query}'.\n", "yellow")
    return all_results, variants

def evaluate_candidates(search_results, core_query, cleaned_query, variants, api_key, log_func, stop_event, use_cache, original_address='', pause_event=None, max_prev_check=3):
    candidates = []
    query_words = core_query.split()
    num_query_words = len(query_words)
    has_digits = re.search(r'^\d', core_query) is not None
    threshold = 0.75 if has_digits else 0.9
    prefix_str = ''
    prefix_norm = ''
    prefix_words = None
    remaining_query = core_query
    if has_digits:
        match = re.match(r'^([\d\-/&\s,]+)', core_query)
        if match:
            prefix_str = match.group(1).strip()
            prefix_norm = re.sub(r'[-/&\s,]+', ' ', prefix_str).strip()
            prefix_words = prefix_norm.split()
            remaining_query = core_query[len(prefix_str):].strip()
    for result in search_results:
        company_name = result.get('title', '')
        core_found = core_name(company_name)
        remaining_found = core_found
        if prefix_words:
            found_match = re.match(r'^([\d\-/&\s,]+)', core_found)
            if found_match:
                found_prefix_str = found_match.group(1).strip()
                found_prefix_norm = re.sub(r'[-/&\s,]+', ' ', found_prefix_str).strip()
                found_prefix_words = found_prefix_norm.split()
                if len(found_prefix_words) < len(prefix_words) or found_prefix_words != prefix_words:
                    continue
                remaining_found = core_found[len(found_prefix_str):].strip()
            else:
                continue
        if not remaining_found and remaining_query:
            continue
        name_sim = calculate_similarity(remaining_query, remaining_found)
        is_exact_match = False
        core_found_words = core_found.split()
        if core_found == core_query or company_name.upper() == cleaned_query or any(company_name.upper() == v for v in variants):
            is_exact_match = True
            name_sim = 1.0
        elif num_query_words >= 2 and ' '.join(core_found_words[:2]) == ' '.join(query_words[:2]):
            name_sim = calculate_similarity(remaining_query, remaining_found)
            if name_sim >= threshold:
                is_exact_match = True

        # Base accuracy from name similarity
        overall_accuracy = name_sim * 100

        # Reward aligned leading tokens (same words in same positions),
        # which helps prefer 'RESOURCE FINDER GROUP' over 'RF RESOURCE FINDER'.
        max_prefix = min(3, num_query_words, len(core_found_words))
        aligned = 0
        for i in range(max_prefix):
            if core_found_words[i] == query_words[i]:
                aligned += 1
            else:
                break
        if aligned > 0:
            overall_accuracy += PREFIX_ALIGNMENT_BONUS_PER_WORD * aligned

        # Reward or penalise based on presence/absence of distinctive words.
        query_tokens = [w for w in query_words if w and not w.isdigit()]
        found_tokens = set(core_found_words)
        missing_distinct = 0
        for token in query_tokens:
            # Ignore very generic company words (already in COMPANY_KEYWORDS)
            if token in COMPANY_KEYWORDS:
                continue
            if token not in found_tokens:
                missing_distinct += 1
        if missing_distinct == 0 and query_tokens:
            # All distinctive words present => small bonus
            overall_accuracy += ALL_DISTINCT_PRESENT_BONUS
        elif missing_distinct > 0:
            # Penalise when distinctive words (like 'POL' or 'CAVERSWALL') are missing
            overall_accuracy -= DISTINCT_MISSING_PENALTY_PER_WORD * missing_distinct
            # If the very first distinctive token is missing, apply an extra penalty,
            # because it usually carries a lot of meaning (e.g. a place or brand).
            first_distinct = next((t for t in query_tokens if t not in COMPANY_KEYWORDS), None)
            if first_distinct and first_distinct not in found_tokens:
                overall_accuracy -= FIRST_DISTINCT_EXTRA_PENALTY

        if overall_accuracy >= MIN_OVERALL_ACCURACY or is_exact_match:
            profile = get_company_profile(result['company_number'], api_key, log_func, stop_event, use_cache)
            if profile:
                status = profile.get('company_status', '')
                creation_date = profile.get('date_of_creation', '1900-01-01')
                if status != 'active':
                    overall_accuracy -= NON_ACTIVE_STATUS_PENALTY
                # Older companies get a small penalty (less likely to be recent debtors)
                try:
                    if datetime.datetime.strptime(creation_date, '%Y-%m-%d').year < datetime.date.today().year - OLD_COMPANY_YEARS:
                        overall_accuracy -= OLD_COMPANY_PENALTY
                except Exception:
                    pass

                # Address / postcode weighting
                if original_address:
                    sheet_pc = extract_postcode(original_address)
                    ro_addr = profile.get('registered_office_address', {}) or {}
                    ro_pc_raw = ro_addr.get('postal_code', '')
                    ro_pc = extract_postcode(ro_pc_raw)
                    if sheet_pc and ro_pc:
                        if sheet_pc == ro_pc:
                            # Strong postcode match: boost score
                            overall_accuracy += POSTCODE_MATCH_BONUS
                            if log_func:
                                log_func(
                                    f"Postcode match boost for {result.get('title','')} "
                                    f"(CN: {result.get('company_number','')}): sheet_pc={sheet_pc}, ro_pc={ro_pc}\n",
                                    "green"
                                )
                        else:
                            # Different postcode: small penalty
                            overall_accuracy -= POSTCODE_MISMATCH_PENALTY
                            if log_func:
                                log_func(
                                    f"Postcode mismatch penalty for {result.get('title','')} "
                                    f"(CN: {result.get('company_number','')}): sheet_pc={sheet_pc}, ro_pc={ro_pc}\n",
                                    "yellow"
                                )
            candidates.append((result, name_sim, overall_accuracy, is_exact_match, False))

    # Optional debug logging of top candidates to help tune thresholds.
    if log_func and candidates:
        sorted_candidates = sorted(candidates, key=lambda x: -x[2])[:3]
        lines = ["Top candidates:"]
        for cand_result, name_sim, overall_accuracy, is_exact_match, is_previous in sorted_candidates:
            lines.append(
                f"  - {cand_result.get('title','')} "
                f"(CN: {cand_result.get('company_number','')}, "
                f"name_sim={int(name_sim*100)}%, overall={int(overall_accuracy)}%, "
                f"exact={is_exact_match}, prev_name={is_previous})"
            )
        log_func("\n".join(lines) + "\n", "debug")

    if not candidates:
        prev_results = search_results[:max_prev_check]
        with ThreadPoolExecutor(max_workers=min(3, len(prev_results))) as executor:
            futures = {}
            for result in prev_results:
                future = executor.submit(get_company_profile, result['company_number'], api_key, log_func, stop_event, use_cache, pause_event)
                futures[future] = result
            profiles = {}
            for future in as_completed(futures):
                if stop_event.is_set():
                    raise Exception("Processing stopped")
                result = futures[future]
                profiles[result['company_number']] = future.result()
        for result in prev_results:
            profile = profiles.get(result['company_number'])
            if profile and 'previous_company_names' in profile:
                for prev in profile['previous_company_names']:
                    prev_name = prev.get('name', '').upper()
                    prev_core = core_name(prev_name)
                    prev_remaining = prev_core
                    if prefix_words:
                        prev_match = re.match(r'^([\d\-/&\s,]+)', prev_core)
                        if prev_match:
                            prev_prefix_str = prev_match.group(1).strip()
                            prev_prefix_norm = re.sub(r'[-/&\s,]+', ' ', prev_prefix_str).strip()
                            prev_prefix_words = prev_prefix_norm.split()
                            if len(prev_prefix_words) < len(prefix_words) or prev_prefix_words != prefix_words:
                                continue
                            prev_remaining = prev_core[len(prev_prefix_str):].strip()
                        else:
                            continue
                    if not prev_remaining and remaining_query:
                        continue
                    name_sim = calculate_similarity(remaining_query, prev_remaining)
                    if name_sim >= 0.85:
                        adjusted_sim = name_sim - 0.05
                        candidates.append((result, adjusted_sim, adjusted_sim * 100, False, True))
                        break
    return candidates

def process_best_match(best_match, api_key, log_func, best_name_sim, best_accuracy, best_is_exact, is_previous, stop_event, use_cache, pause_event=None):
    company_number = best_match.get('company_number', '')
    profile = get_company_profile(company_number, api_key, log_func, stop_event, use_cache, pause_event)
    proposed_address = ''
    status = ''
    status_detail = ''
    if profile:
        ro_address = profile.get('registered_office_address', {})
        proposed_address = ', '.join([ro_address.get(k, '') for k in ['address_line_1', 'address_line_2', 'locality', 'postal_code'] if ro_address.get(k)])
        status = profile.get('company_status', '')
        status_detail = profile.get('company_status_detail', '')
    officers = get_officers(company_number, api_key, log_func, stop_event, use_cache, pause_event)
    directors_names = []
    directors_addresses = []
    disq_directors = []
    with ThreadPoolExecutor(max_workers=min(3, len(officers))) as executor:
        futures = [executor.submit(get_disqualified_info, officer.get('name', '').title() if ', ' not in officer.get('name', '') else f"{officer.get('name', '').split(', ', 1)[1].capitalize()} {officer.get('name', '').split(', ', 1)[0].capitalize()}", api_key, log_func, stop_event, use_cache, pause_event) for officer in officers if 'resigned_on' not in officer and 'director' in officer.get('officer_role', '')]
        disq_results = [future.result() for future in as_completed(futures)]
    for officer, (is_dq, details) in zip([o for o in officers if 'resigned_on' not in o and 'director' in o.get('officer_role', '')], disq_results):
        if stop_event.is_set():
            raise Exception("Processing stopped")
        name = officer.get('name', '')
        formatted_name = name.title() if ', ' not in name else f"{name.split(', ', 1)[1].capitalize()} {name.split(', ', 1)[0].capitalize()}"
        addr = officer.get('address', {})
        director_addr = ', '.join([addr.get(k, '') for k in ['premises', 'address_line_1', 'locality', 'postal_code'] if addr.get(k)])
        directors_names.append(formatted_name)
        directors_addresses.append(director_addr)
        if is_dq:
            disq_directors.append(f"{formatted_name} ({details})")
    directors_names_str = '; '.join(directors_names)
    directors_addresses_str = '; '.join(directors_addresses)
    disq_directors_str = '; '.join(disq_directors)
    insolvency_type, insolvency_date = get_insolvency(company_number, api_key, log_func, stop_event, use_cache, pause_event)
    if status == 'dissolved' and not insolvency_type:
        insolvency_type = 'dissolved'
        insolvency_date = profile.get('date_of_cessation', '')
    if not insolvency_type and status_detail == 'active-proposal-to-strike-off':
        insolvency_type = 'ptso'
        filings = get_filing_history(company_number, api_key, log_func, stop_event, use_cache, pause_event)
        for filing in filings:
            if stop_event.is_set():
                raise Exception("Processing stopped")
            filing_type = filing.get('type', '')
            if filing_type in ['DS01', 'GAZ1']:
                insolvency_date = filing.get('action_date', '') or filing.get('date', '')
                break
    if insolvency_type:
        status = insolvency_types.get(insolvency_type, insolvency_type)
    explanation = f"Companies House - Name sim: {int(best_name_sim * 100)}%"
    if best_is_exact:
        explanation += " (Exact name match)"
    if is_previous:
        explanation += " (Previous name match)"
    return {
        'Found Company Name': best_match.get('title', ''),
        'Company Number': company_number,
        'Proposed New Address': proposed_address,
        'Insolvency Status': status,
        'Insolvency Type': insolvency_type,
        'Date of Appointment': insolvency_date,
        'Directors Names': directors_names_str,
        'Directors Addresses': directors_addresses_str,
        'Disqualified Directors': disq_directors_str,
        'Match Accuracy': f"{int(best_accuracy)}%",
        'Match Explanation': explanation
    }

def process_row(row, api_key, log_func, name_column, stop_event, use_cache, pause_event=None):
    start_time = time.time()
    debtor_name = row[name_column]
    cleaned = clean_name(debtor_name)
    if not is_likely_company(cleaned):
        updates = {'Found Company Name': '', 'Company Number': '', 'Proposed New Address': '', 'Insolvency Status': '', 'Insolvency Type': '', 'Date of Appointment': '', 'Directors Names': '', 'Directors Addresses': '', 'Disqualified Directors': '', 'Match Accuracy': '0%', 'Match Explanation': 'Not a company'}
        new_row = row.to_dict()
        new_row.update(updates)
        results_list = [new_row]
        duration = time.time() - start_time
        return results_list, duration
    # Build original address string early so we can use it for scoring.
    original_address_parts = [
        str(row.get(key, ''))
        for key in [
            'Address Name or Number',
            'Address Line1',
            'Address Line2',
            'Address Line3',
            'Address Line4',
            'Add Post Code',
        ]
        if key in row and not pd.isna(row[key])
    ]
    original_address = ', '.join(original_address_parts).upper()

    cleaned_query, core_query, search_query = prepare_query(debtor_name)
    search_results, variants = search_and_variants(search_query, cleaned_query, api_key, log_func, stop_event, use_cache, pause_event=pause_event)
    candidates = evaluate_candidates(search_results, core_query, cleaned_query, variants, api_key, log_func, stop_event, use_cache, original_address=original_address, pause_event=pause_event)
    results_list = []
    if candidates:
        candidates.sort(key=lambda x: - (x[1] * min(1.0, len(x[0]['title']) / len(cleaned_query))))
        best_match, best_name_sim, best_accuracy, best_is_exact, is_previous = candidates[0]
        original_address = original_address
        profile = get_company_profile(best_match['company_number'], api_key, log_func, stop_event, use_cache, pause_event)
        addr_sim = None
        if profile and original_address:
            ro_address = ', '.join([profile.get('registered_office_address', {}).get(k, '').upper() for k in ['address_line_1', 'address_line_2', 'locality', 'postal_code']])
            addr_sim = fuzz.partial_ratio(original_address, ro_address) / 100.0
            # If address is quite different, reduce overall accuracy but do not hard reject.
            if addr_sim < ADDRESS_LOW_SIM_THRESHOLD:
                best_accuracy -= ADDRESS_LOW_SIM_PENALTY
        updates = process_best_match(best_match, api_key, log_func, best_name_sim, best_accuracy, best_is_exact, is_previous, stop_event, use_cache, pause_event=pause_event)
        if addr_sim is not None:
            addr_percent = int(addr_sim * 100)
            updates['Match Explanation'] += f" | Address sim: {addr_percent}%"
            if addr_sim < 0.7:
                updates['Match Explanation'] += " (Low address similarity, accuracy reduced)"
        new_row = row.to_dict()
        new_row.update(updates)
        results_list.append(new_row)
    else:
        updates = {'Found Company Name': '', 'Company Number': '', 'Proposed New Address': '', 'Insolvency Status': '', 'Insolvency Type': '', 'Date of Appointment': '', 'Directors Names': '', 'Directors Addresses': '', 'Disqualified Directors': '', 'Match Accuracy': '0%', 'Match Explanation': 'No match found'}
        new_row = row.to_dict()
        new_row.update(updates)
        results_list.append(new_row)
    duration = time.time() - start_time
    return results_list, duration

def update_log(text, tag=None):
    def do_update():
        # Suppress debug logs unless verbose mode is enabled in the GUI.
        try:
            if tag == "debug" and verbose_var is not None and not verbose_var.get():
                return
        except Exception:
            # If verbose_var is not yet initialised, fall back to showing the log.
            pass
        active_log_text.insert(tk.END, text, tag)
        active_log_text.see(tk.END)
    active_log_text.after(0, do_update)

def init_log_context_menu(widget):
    """Attach a right-click context menu with copy/select/clear to a text widget."""
    menu = tk.Menu(widget, tearoff=0)

    def copy_selection():
        try:
            text = widget.get("sel.first", "sel.last")
        except tk.TclError:
            return
        if text:
            widget.clipboard_clear()
            widget.clipboard_append(text)

    def select_all():
        widget.tag_add("sel", "1.0", "end-1c")

    def clear_log():
        widget.delete("1.0", tk.END)

    menu.add_command(label="Copy", command=copy_selection)
    menu.add_command(label="Select All", command=select_all)
    menu.add_separator()
    menu.add_command(label="Clear Log", command=clear_log)

    def show_menu(event):
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    widget.bind("<Button-3>", show_menu)

def show_message(title, message):
    def do_show():
        messagebox.showinfo(title, message)
    root.after(0, do_show)

def show_about():
    """Display a simple About dialog with version and contact info."""
    info = (
        f"{APP_NAME} v{APP_VERSION}\n\n"
        "Bulk Companies House lookup tool for Excel spreadsheets.\n\n"
        "Developer: Angus Mackay\n"
        "Contact: Angus_Mackay_@hotmail.com"
    )
    messagebox.showinfo("About", info)

def save_results(results, output_file, name_column, selected_columns, graph_selections):
    full_df = pd.DataFrame(results)
    df_output = full_df.copy()
    df_output = df_output.dropna(axis=1, how='all')
    cols = list(df_output.columns)
    if 'Original Address' in cols and name_column in cols:
        cols.insert(cols.index(name_column) + 1, cols.pop(cols.index('Original Address')))
    if selected_columns:
        existing_cols = [c for c in selected_columns if c in df_output.columns]
        df_output = df_output[existing_cols]
    df_output.to_excel(output_file, index=False)
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    balance_col_idx = None
    headers = [cell.value for cell in ws[1]]
    tel_col_idx = headers.index('Tel1Number') + 1 if 'Tel1Number' in headers else None
    for col_idx in range(1, ws.max_column + 1):
        header = headers[col_idx - 1]
        if header and ('balance' in header.lower() or 'amount' in header.lower()):
            balance_col_idx = col_idx
        if header and 'date' in header.lower():
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        date_val = datetime.datetime.strptime(str(cell.value), '%Y-%m-%d')
                        cell.value = date_val
                        cell.number_format = 'dd/mm/yy'
                    except ValueError:
                        pass
    if balance_col_idx:
        for row_idx in range(2, ws.max_row +1):
            ws.cell(row=row_idx, column=balance_col_idx).number_format = '"£"#,##0.00'
    if tel_col_idx:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=tel_col_idx)
            if cell.value:
                tel_str = str(cell.value)
                if len(tel_str) == 10 and tel_str[0] == '7':
                    cell.value = '0' + tel_str
                    cell.number_format = '@'

    fills = []
    for _, row in full_df.iterrows():
        status = row.get('Insolvency Status', '') or ''
        ins_type = row.get('Insolvency Type', '') or ''
        disq_dir = row.get('Disqualified Directors', '') or ''
        match_acc = row.get('Match Accuracy', '0%') or '0%'
        status_lower = status.lower()
        is_insolvent = bool(ins_type) or any(term in status_lower for term in ['proposal to strike off', 'dissolved', 'liquidation', 'administration', 'receivership', 'insolvency', 'voluntary arrangement']) or bool(disq_dir)
        is_match = match_acc != '0%'
        fill = red_fill if is_insolvent else green_fill if is_match else None
        fills.append(fill)

    for row_idx in range(2, ws.max_row + 1):
        fill = fills[row_idx - 2]
        if fill:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    summary_ws = wb.create_sheet(title='Summary')
    balance_col = next((col for col in full_df.columns if 'balance' in col.lower() or 'amount' in col.lower()), None)
    row_offset = 1
    chart_offset = "E1"
    if 'Insolvent vs Active' in graph_selections:
        if not full_df.empty and 'Match Accuracy' in full_df.columns:
            df_matched = full_df[full_df['Match Accuracy'] != '0%']
        else:
            df_matched = pd.DataFrame()
        if not df_matched.empty:
            df_insolvent = df_matched.apply(lambda row: bool(row['Insolvency Type']) or any(term in str(row['Insolvency Status']).lower() for term in ['proposal to strike off', 'dissolved', 'liquidation', 'administration', 'receivership', 'insolvency', 'voluntary arrangement']) or bool(row['Disqualified Directors']), axis=1)
            insolvent_count = df_insolvent.sum()
            active_count = len(df_matched) - insolvent_count
            insolvent_value = df_matched[df_insolvent][balance_col].sum() if balance_col else 0
            active_value = df_matched[~df_insolvent][balance_col].sum() if balance_col else 0
        else:
            insolvent_count = active_count = insolvent_value = active_value = 0

        summary_ws.cell(row=row_offset, column=1, value='Category')
        summary_ws.cell(row=row_offset, column=2, value='Count')
        summary_ws.cell(row=row_offset, column=3, value='Value')
        summary_ws.cell(row=row_offset+1, column=1, value='Insolvent')
        summary_ws.cell(row=row_offset+1, column=2, value=insolvent_count)
        summary_ws.cell(row=row_offset+1, column=3, value=insolvent_value)
        summary_ws.cell(row=row_offset+2, column=1, value='Active')
        summary_ws.cell(row=row_offset+2, column=2, value=active_count)
        summary_ws.cell(row=row_offset+2, column=3, value=active_value)

        if balance_col:
            summary_ws.cell(row=row_offset+1, column=3).number_format = '"£"#,##0.00'
            summary_ws.cell(row=row_offset+2, column=3).number_format = '"£"#,##0.00'

        pie_count = PieChart()
        labels = Reference(summary_ws, min_col=1, min_row=row_offset+1, max_row=row_offset+2)
        data_count = Reference(summary_ws, min_col=2, min_row=row_offset, max_row=row_offset+2)
        pie_count.add_data(data_count, titles_from_data=True)
        pie_count.set_categories(labels)
        pie_count.title = "Insolvencies vs Active by Count"
        summary_ws.add_chart(pie_count, chart_offset)

        pie_value = PieChart()
        data_value = Reference(summary_ws, min_col=3, min_row=row_offset, max_row=row_offset+2)
        pie_value.add_data(data_value, titles_from_data=True)
        pie_value.set_categories(labels)
        pie_value.title = "Insolvencies vs Active by Value"
        summary_ws.add_chart(pie_value, f"E{row_offset+19}")

        row_offset += 4
        chart_offset = f"E{row_offset}"

    wb.save(output_file)

def compute_summary_stats(results):
    """Compute simple summary stats from the results list."""
    if not results:
        return 0, 0, 0
    full_df = pd.DataFrame(results)
    total = len(full_df)
    matched = 0
    insolvent = 0
    if 'Match Accuracy' in full_df.columns:
        matched = (full_df['Match Accuracy'] != '0%').sum()
    if 'Insolvency Type' in full_df.columns or 'Insolvency Status' in full_df.columns or 'Disqualified Directors' in full_df.columns:
        def is_ins(row):
            status = str(row.get('Insolvency Status', '') or '').lower()
            ins_type = str(row.get('Insolvency Type', '') or '')
            disq = str(row.get('Disqualified Directors', '') or '')
            return bool(ins_type) or any(term in status for term in [
                'proposal to strike off', 'dissolved', 'liquidation',
                'administration', 'receivership', 'insolvency', 'voluntary arrangement'
            ]) or bool(disq)
        insolvent = full_df.apply(is_ins, axis=1).sum()
    return int(total), int(matched), int(insolvent)

def process_file(file_path, start_row, end_row, api_key, output_file, pause_event, stop_event, progress_bar, status_label, name_column, selected_columns, graph_selections, use_cache):
    """
    GUI processing: use *fresh* in-memory caches for each app session.
    We do not load or save pickle cache files here; caching is only for the
    lifetime of the running app to ensure results are always fresh per launch.
    """
    global profile_cache, officers_cache, insolvency_cache, filing_cache, disqualified_cache, search_cache
    # If the user has turned off cache, clear any in-memory state; otherwise
    # keep whatever has been built during this GUI session.
    if not use_cache:
        profile_cache.clear()
        officers_cache.clear()
        insolvency_cache.clear()
        filing_cache.clear()
        disqualified_cache.clear()
        search_cache.clear()
    results = []
    try:
        df_input = pd.read_excel(file_path, sheet_name=0)
    except Exception as e:
        update_log(f"Failed to read file: {str(e)}\n", "red")
        return
    if name_column not in df_input.columns:
        update_log(f"No '{name_column}' column found.\n", "red")
        return
    row_start = max(0, start_row - 1)
    row_end = min(len(df_input), end_row)
    row_count = row_end - row_start
    total_start_time = time.time()
    pause_start = None
    save_interval = 100
    times = []
    prev_api_calls = 0

    def log_func(text, tag=None):
        update_log(text, tag)

    progress_bar['maximum'] = row_count
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {}
        for rel_index, abs_index in enumerate(range(row_start, row_end), 1):
            row = df_input.iloc[abs_index]
            debtor_name = row[name_column]
            abs_row_num = abs_index + 2
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            update_log(f"{current_time} - Submitted row {abs_row_num} ({rel_index}/{row_count}): {debtor_name}\n")
            future = executor.submit(process_row, row, api_key, log_func, name_column, stop_event, use_cache, pause_event=pause_event)
            futures[future] = (rel_index, abs_row_num, debtor_name)

        processed = 0
        for future in as_completed(futures):
            if stop_event.is_set():
                for f in futures:
                    f.cancel()
                update_log(f"Processing stopped after {time.time() - total_start_time:.2f} seconds.\n")
                break
            while pause_event.is_set():
                # Allow stop to take effect even while paused.
                if stop_event.is_set():
                    break
                if pause_start is None:
                    pause_start = time.time()
                    update_log("Processing paused.\n", "yellow")
                time.sleep(0.1)
            if stop_event.is_set():
                for f in futures:
                    f.cancel()
                update_log(f"Processing stopped after {time.time() - total_start_time:.2f} seconds.\n")
                break
            if pause_start is not None:
                pause_duration = time.time() - pause_start
                update_log(f"Processing resumed after {pause_duration:.2f} seconds pause.\n", "green")
                pause_start = None
            rel_index, abs_row_num, debtor_name = futures[future]
            try:
                results_list, duration = future.result()
            except Exception as e:
                if str(e) == "Processing stopped":
                    break
                update_log(f"Error processing row {abs_row_num}: {str(e)}\n", "red")
                results_list = []
                duration = 0.0
            times.append(duration)
            for new_row in results_list:
                original_address_parts = [str(new_row.get(key, '')) for key in ['Address Name or Number', 'Address Line1', 'Address Line2', 'Address Line3', 'Address Line4', 'Add Post Code'] if key in new_row and not pd.isna(new_row[key])]
                new_row['Original Address'] = ', '.join(original_address_parts)
                results.append(new_row)
            accuracy = results_list[0].get('Match Accuracy', '0%') if results_list else '0%'
            matched_name = results_list[0].get('Found Company Name', '') if results_list and accuracy != '0%' else ''
            explanation = results_list[0].get('Match Explanation', '') if results_list else ''
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if accuracy != '0%':
                acc_num = int(accuracy.rstrip('%'))
                tag = "green" if acc_num > 80 else "yellow"
                company_no = results_list[0].get('Company Number', '')
                status = results_list[0].get('Insolvency Status', '')
                explanation = results_list[0].get('Match Explanation', '')
                update_log(
                    f"{current_time} - Searched: {debtor_name} - Matched: {matched_name} "
                    f"(CN: {company_no}, Status: {status}) with accuracy {accuracy} "
                    f"- Duration: {duration:.2f} seconds | {explanation}\n",
                    tag
                )
            elif explanation == 'Not a company':
                update_log(f"{current_time} - {debtor_name} - Not a company - Duration: {duration:.2f} seconds\n", "yellow")
            else:
                update_log(f"{current_time} - {debtor_name} - Not Found - Duration: {duration:.2f} seconds\n", "red")
            processed += 1
            progress_bar['value'] = processed
            if times:
                avg_time = sum(times) / len(times)
                remaining = row_count - processed
                eta = avg_time * remaining
                elapsed = time.time() - total_start_time
                pauses_total = rate_limiter.pauses
                remaining_calls = rate_limiter.max_requests - len(rate_limiter.requests)
                avg_calls_per_row = api_calls / processed if processed > 0 else 0
                avg_api_time = sum(api_times) / len(api_times) if api_times else 0
                status_text = f"Elapsed: {elapsed//60:.0f}m {elapsed%60:.0f}s | ETA: {eta//60:.0f}m {eta%60:.0f}s | API Calls: {api_calls} | Remaining Calls: {remaining_calls} | Pauses: {pauses_total} | Avg Calls/Row: {avg_calls_per_row:.1f} | Avg API Time: {avg_api_time:.2f}s"
                def update_status():
                    status_label.config(text=status_text)
                root.after(0, update_status)
            root.update_idletasks()
            time.sleep(0.1)
            prev_api_calls = api_calls
            if processed % save_interval == 0:
                # Intermediate Excel save only (no disk cache in GUI mode).
                try:
                    save_results(results, output_file, name_column, selected_columns, graph_selections)
                    update_log(f"Intermediate save at row {abs_row_num}\n", "green")
                except Exception as e:
                    update_log(f"Error saving intermediate results to {output_file}: {e}\n", "red")
                    show_message("Save Error", f"Error saving intermediate results:\n{e}")
                    break
    # Final Excel save only (no disk cache in GUI mode).
    try:
        save_results(results, output_file, name_column, selected_columns, graph_selections)
    except Exception as e:
        update_log(f"Error saving final results to {output_file}: {e}\n", "red")
        show_message("Save Error", f"Error saving final results:\n{e}")
        enable_widgets()
        progress_bar.destroy()
        return

    total, matched, insolvent = compute_summary_stats(results)
    unmatched = total - matched
    summary_text = (
        f"Done! Results saved to {output_file}\n"
        f"Total rows processed: {total}\n"
        f"Matched companies: {matched}\n"
        f"Insolvent / flagged: {insolvent}\n"
        f"Unmatched / non-companies: {unmatched}"
    )
    update_log(summary_text + "\n", "green")
    show_message("Processing complete", summary_text)
    enable_widgets()
    progress_bar.destroy()

def set_widget_state(widget, state, exceptions):
    if widget in exceptions:
        return
    try:
        widget.config(state=state)
    except:
        pass
    for child in widget.winfo_children():
        set_widget_state(child, state, exceptions)

def disable_widgets():
    # Do not disable log widgets so logging continues during processing.
    exceptions = [pause_button, stop_button, toggle_log_button, detach_button, ch_link_label]
    # Keep verbose toggle active so logging detail can be changed mid-run.
    try:
        exceptions.append(verbose_check_button)
    except NameError:
        pass
    try:
        exceptions.append(embedded_log_text)
    except NameError:
        pass
    try:
        if detached_log_text is not None:
            exceptions.append(detached_log_text)
    except NameError:
        pass
    set_widget_state(root, 'disabled', exceptions)

def enable_widgets():
    set_widget_state(root, 'normal', [])

custom_api_key = ''

def on_ch_mode_change(*args):
    global custom_api_key
    mode = ch_mode_var.get()
    if mode == "Gus's API":
        # Remember whatever was in the box as the custom key (if not Gus's key already)
        current = api_key_entry.get().strip()
        if current and current != GUS_API_KEY:
            custom_api_key = current
        api_key_entry.delete(0, tk.END)
        api_key_entry.insert(0, GUS_API_KEY)
    elif mode == "Blank":
        # Store any non-blank, non-Gus key as custom, then clear
        current = api_key_entry.get().strip()
        if current and current not in ("", GUS_API_KEY):
            custom_api_key = current
        api_key_entry.delete(0, tk.END)
    else:  # Custom
        api_key_entry.delete(0, tk.END)
        api_key_entry.insert(0, custom_api_key if custom_api_key else '')

def start_processing():
    file_path = file_var.get()
    if not file_path:
        messagebox.showerror("Error", "Please select a file.")
        return
    output_file = output_var.get()
    if not output_file:
        messagebox.showerror("Error", "Please specify output file.")
        return
    try:
        start_row = int(start_row_entry.get())
        end_row = int(end_row_entry.get())
    except ValueError:
        messagebox.showerror("Error", "Invalid row range.")
        return
    api_key = api_key_entry.get().strip()
    if not api_key:
        messagebox.showerror("Error", "API key is required.")
        return
    name_column = name_column_var.get().strip()
    if not name_column:
        messagebox.showerror("Error", "Name column is required.")
        return
    selected_columns = list(selected_listbox.get(0, END))
    graph_selections = []
    if insolvent_var.get():
        graph_selections.append('Insolvent vs Active')
    use_cache = use_cache_var.get()
    pause_event = threading.Event()
    stop_event = threading.Event()
    # Progress bar and status label live inside the status bar frame.
    progress_bar = ttk.Progressbar(
        status_frame,
        orient='horizontal',
        mode='determinate',
        style='green.Horizontal.TProgressbar',
        length=300,
    )
    progress_bar.grid(row=0, column=0, sticky='ew', padx=4, pady=(2, 0))
    status_label = tk.Label(
        status_frame,
        text="",
        font=('Arial', 11, 'bold'),
        fg='darkblue',
        bg='#e0e0e0',
        anchor='w',
    )
    status_label.grid(row=1, column=0, sticky='ew', padx=4, pady=(0, 2))
    thread = threading.Thread(target=process_file, args=(file_path, start_row, end_row, api_key, output_file, pause_event, stop_event, progress_bar, status_label, name_column, selected_columns, graph_selections, use_cache))
    thread.daemon = True
    thread.start()
    disable_widgets()
    pause_button.config(text="Pause", command=lambda: toggle_pause(pause_event, pause_button))
    stop_button.config(command=stop_event.set)

def toggle_pause(pause_event, button):
    if pause_event.is_set():
        pause_event.clear()
        button.config(text="Pause")
    else:
        pause_event.set()
        button.config(text="Resume")

def load_spreadsheet(file_path):
    """Load spreadsheet metadata into the GUI from a given path."""
    if not file_path:
        return
    file_var.set(file_path)
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        total_rows = len(df)
        start_row_entry.delete(0, tk.END)
        start_row_entry.insert(0, "1")
        end_row_entry.delete(0, tk.END)
        end_row_entry.insert(0, str(total_rows))
        update_log(f"Loaded spreadsheet with {total_rows} rows. Range set to 1-{total_rows}.\n", "green")
        columns = list(df.columns)
        default_name = 'Debtor Name' if 'Debtor Name' in columns else columns[0] if columns else ""
        name_column_var.set(default_name)
        menu = name_column_menu["menu"]
        menu.delete(0, "end")
        for col in columns:
            menu.add_command(label=col, command=lambda value=col: name_column_var.set(value))
        available_listbox.delete(0, END)
        for col in columns + ['Found Company Name', 'Company Number', 'Proposed New Address', 'Insolvency Status', 'Insolvency Type', 'Date of Appointment', 'Directors Names', 'Directors Addresses', 'Disqualified Directors', 'Match Accuracy', 'Match Explanation', 'Original Address']:
            available_listbox.insert(END, col)
        load_profile()
        update_est()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file info: {e}")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        load_spreadsheet(file_path)
        # Persist the last opened spreadsheet path.
        current_settings = load_settings()
        current_settings['spreadsheet'] = file_path
        save_settings(current_settings)

def browse_output():
    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if output_path:
        output_var.set(output_path)
        # Persist the last chosen output file.
        current_settings = load_settings()
        current_settings['output_file'] = output_path
        save_settings(current_settings)

def toggle_log():
    if is_detached:
        if log_window.winfo_viewable():
            log_window.withdraw()
        else:
            log_window.deiconify()
    else:
        if embedded_log_text.winfo_ismapped():
            embedded_log_text.grid_remove()
        else:
            embedded_log_text.grid()

def toggle_detach():
    global is_detached, active_log_text, log_window, detached_log_text
    if is_detached:
        current_text = active_log_text.get("1.0", END)
        log_window.destroy()
        detached_log_text = None
        embedded_log_text.delete("1.0", END)
        embedded_log_text.insert(END, current_text)
        embedded_log_text.grid(row=7, column=0, columnspan=4, sticky='nsew')
        active_log_text = embedded_log_text
        is_detached = False
        detach_button.config(text="Detach Log", command=toggle_detach)
    else:
        current_text = active_log_text.get("1.0", END)
        embedded_log_text.grid_remove()
        log_window = tk.Toplevel(root)
        log_window.title("Log")
        log_window.geometry("800x400")
        log_window.rowconfigure(0, weight=1)
        log_window.columnconfigure(0, weight=1)
        detached_log_text = scrolledtext.ScrolledText(log_window, wrap='none', bg="white", fg="black")
        detached_log_text.grid(row=0, column=0, sticky='nsew')
        detached_log_text.insert(END, current_text)
        detached_log_text.tag_configure("green", foreground="green")
        detached_log_text.tag_configure("red", foreground="red")
        detached_log_text.tag_configure("yellow", foreground="orange")
        init_log_context_menu(detached_log_text)
        active_log_text = detached_log_text
        is_detached = True
        detach_button.config(text="Attach Log", command=toggle_detach)
        log_window.protocol("WM_DELETE_WINDOW", toggle_detach)

def add_column():
    selected = available_listbox.curselection()
    for i in selected:
        col = available_listbox.get(i)
        if col not in selected_listbox.get(0, END):
            selected_listbox.insert(END, col)

def remove_column():
    selected = selected_listbox.curselection()
    for i in sorted(selected, reverse=True):
        selected_listbox.delete(i)

def move_up():
    selected = selected_listbox.curselection()
    if not selected:
        return
    for i in selected:
        if i == 0:
            continue
        text = selected_listbox.get(i)
        selected_listbox.delete(i)
        selected_listbox.insert(i-1, text)
        selected_listbox.select_set(i-1)

def move_down():
    selected = selected_listbox.curselection()
    if not selected:
        return
    for i in sorted(selected, reverse=True):
        if i == selected_listbox.size() - 1:
            continue
        text = selected_listbox.get(i)
        selected_listbox.delete(i)
        selected_listbox.insert(i+1, text)
        selected_listbox.select_set(i+1)

def select_all_columns():
    selected_listbox.delete(0, END)
    for i in range(available_listbox.size()):
        col = available_listbox.get(i)
        selected_listbox.insert(END, col)

def save_profile():
    current_profile = profile_var.get()
    selected = list(selected_listbox.get(0, END))
    profiles = load_profiles()
    if current_profile == "Default":
        profile_name = simpledialog.askstring("Save Profile", "Enter profile name:")
        if not profile_name:
            return
        if profile_name == "Default":
            messagebox.showerror("Error", "Cannot use Default as profile name.")
            return
    else:
        profile_name = current_profile
    profiles[profile_name] = selected
    with open('profiles.json', 'w') as f:
        json.dump(profiles, f)
    load_profile_menu()
    profile_var.set(profile_name)
    load_profile()

def delete_profile():
    profile_name = profile_var.get()
    if profile_name == "Default":
        messagebox.showerror("Error", "Cannot delete default profile.")
        return
    profiles = load_profiles()
    if profile_name in profiles:
        del profiles[profile_name]
        with open('profiles.json', 'w') as f:
            json.dump(profiles, f)
        load_profile_menu()
        profile_var.set("Default")
        load_profile()

def load_profiles():
    if os.path.exists('profiles.json'):
        with open('profiles.json', 'r') as f:
            profiles = json.load(f)
            if 'Default' in profiles:
                del profiles['Default']
            return profiles
    return {}

def load_profile(*args):
    profile_name = profile_var.get()
    selected_listbox.delete(0, END)
    if profile_name == "Default":
        select_all_columns()
        return
    profiles = load_profiles()
    if profile_name in profiles:
        for col in profiles[profile_name]:
            if col in list(available_listbox.get(0, END)):
                selected_listbox.insert(END, col)

def load_profile_menu():
    profiles = load_profiles()
    menu = profile_menu["menu"]
    menu.delete(0, "end")
    menu.add_command(label="Default", command=lambda: profile_var.set("Default"))
    for p in sorted(profiles.keys()):
        menu.add_command(label=p, command=lambda value=p: profile_var.set(value))

def update_est():
    """
    Estimate total time based on selected row range and avg calls/row,
    respecting the Companies House guideline of ~600 calls per 5 minutes.
    """
    try:
        start = int(start_row_entry.get())
        end = int(end_row_entry.get())
        rows = max(0, end - start + 1)
        avg = avg_calls.get()
        total_calls = rows * avg
        # Companies House free tier guideline: 600 requests per 5 minutes.
        calls_per_window = RATE_LIMIT_MAX
        window_seconds = RATE_LIMIT_WINDOW
        time_sec = (total_calls / calls_per_window) * window_seconds
        time_min = time_sec / 60
        time_hours = time_min / 60
        est_label.config(
            text=(
                f"Est. time: {time_min:.1f} min (~{time_hours:.2f} h) "
                f"(total API calls ≈ {int(total_calls)})"
            )
        )
    except ValueError:
        est_label.config(text="Invalid range")

def update_est_event(event):
    update_est()

def persist_runtime_settings():
    """Persist current GUI settings to disk (called on close)."""
    try:
        current = load_settings()
        current['spreadsheet'] = file_var.get()
        current['output_file'] = output_var.get()
        current['name_column'] = name_column_var.get()
        current['ch_mode'] = ch_mode_var.get()
        current['profile'] = profile_var.get()
        current['use_cache'] = bool(use_cache_var.get())
        current['verbose_log'] = bool(verbose_var.get())
        current['window_geometry'] = root.winfo_geometry()
        save_settings(current)
    except Exception:
        pass

def create_gui():
    global root, file_var, output_var, name_column_var, ch_mode_var, insolvent_var, profile_var, use_cache_var
    global start_row_entry, end_row_entry, name_column_menu, avg_calls, est_label
    global api_key_entry, ch_link_label, available_listbox, selected_listbox, profile_menu
    global embedded_log_text, active_log_text, is_detached, log_window, detached_log_text
    global pause_button, stop_button, toggle_log_button, detach_button, status_frame, verbose_check_button

    root = tk.Tk()
    root.title(f"{APP_NAME} v{APP_VERSION}")
    # Start slimmer but taller so the profile/columns and buttons are fully visible,
    # and the log only takes roughly the lower half by default.
    # You can freely resize it; no minimum size is enforced.
    root.geometry("760x860")
    root.resizable(True, True)
    root.columnconfigure(0, weight=1)
    # Give significantly more vertical space to the columns/profile area than the log.
    # This makes the log roughly half the height it was previously.
    root.rowconfigure(4, weight=5)  # columns frame and related controls
    root.rowconfigure(7, weight=1)  # log
    root.configure(padx=10, pady=10, bg="#f0f0f0")

    # Load persisted settings (if any) before creating widgets.
    settings = load_settings()

    # Menu bar: File (Exit, Clear Caches) and Help (About).
    menubar = tk.Menu(root)
    file_menu = tk.Menu(menubar, tearoff=0)
    file_menu.add_command(label="Clear Caches", command=lambda: clear_all_caches(delete_files=True))
    file_menu.add_separator()
    file_menu.add_command(label="Exit", command=root.quit)
    menubar.add_cascade(label="File", menu=file_menu)

    help_menu = tk.Menu(menubar, tearoff=0)
    help_menu.add_command(label="About", command=show_about)
    menubar.add_cascade(label="Help", menu=help_menu)

    root.config(menu=menubar)

    style = ttk.Style()
    style.theme_use('clam')
    style.configure('TButton', padding=6, relief="flat", background="#4CAF50", foreground="white")
    style.map('TButton', background=[('active', '#45a049')])
    style.configure('TRed.TButton', background="#f44336", foreground="white")
    style.map('TRed.TButton', background=[('active', '#e53935')])
    style.configure('TLabel', background="#f0f0f0")
    style.configure('TEntry', fieldbackground="white")
    style.configure('TCheckbutton', background="#f0f0f0")
    style.configure('TOptionMenu', background="#f0f0f0")
    style.configure('green.Horizontal.TProgressbar', foreground='green', background='green')
    style.configure('Horizontal.TScale', background="#f0f0f0")

    file_var = tk.StringVar(value=settings.get("spreadsheet", ""))
    output_var = tk.StringVar(value=settings.get("output_file", "insolvency_results.xlsx"))
    name_column_var = tk.StringVar(value=settings.get("name_column", ""))
    ch_mode_var = tk.StringVar(value=settings.get("ch_mode", "Custom"))
    insolvent_var = tk.BooleanVar(value=True)
    profile_var = tk.StringVar(value=settings.get("profile", "Default"))
    use_cache_var = tk.BooleanVar(value=settings.get("use_cache", True))
    # Verbose logging toggle for GUI log view.
    global verbose_var
    verbose_var = tk.BooleanVar(value=settings.get("verbose_log", False))

    input_frame = tk.Frame(root, bg="#f0f0f0")
    input_frame.grid(row=0, column=0, sticky='ew', pady=5)
    input_frame.columnconfigure(1, weight=1)

    ttk.Label(input_frame, text="Spreadsheet:").grid(row=0, column=0, sticky='e')
    spreadsheet_entry = ttk.Entry(input_frame, textvariable=file_var, width=50)
    spreadsheet_entry.grid(row=0, column=1, sticky='ew')
    browse_file_button = ttk.Button(input_frame, text="Browse", command=browse_file)
    browse_file_button.grid(row=0, column=2)

    ttk.Label(input_frame, text="Start Row:").grid(row=1, column=0, sticky='e')
    start_row_entry = ttk.Entry(input_frame)
    start_row_entry.grid(row=1, column=1, sticky='ew')

    ttk.Label(input_frame, text="End Row:").grid(row=2, column=0, sticky='e')
    end_row_entry = ttk.Entry(input_frame)
    end_row_entry.grid(row=2, column=1, sticky='ew')

    ttk.Label(input_frame, text="Name Column:").grid(row=3, column=0, sticky='e')
    name_column_menu = ttk.OptionMenu(input_frame, name_column_var, "")
    name_column_menu.grid(row=3, column=1, sticky='ew')

    ttk.Label(input_frame, text="Avg calls/row:").grid(row=4, column=0, sticky='e')
    avg_calls = ttk.Scale(input_frame, from_=5, to=20, orient='horizontal')
    avg_calls.set(20)
    avg_calls.grid(row=4, column=1, sticky='ew')
    est_label = ttk.Label(input_frame, text="")
    est_label.grid(row=5, column=0, columnspan=3, sticky='ew')

    start_row_entry.bind("<KeyRelease>", update_est_event)
    end_row_entry.bind("<KeyRelease>", update_est_event)
    avg_calls.bind("<ButtonRelease-1>", update_est_event)

    api_frame = tk.Frame(root, bg="#f0f0f0")
    api_frame.grid(row=1, column=0, sticky='ew', pady=5)
    api_frame.columnconfigure(1, weight=1)

    ttk.Label(api_frame, text="Companies House API Mode:").grid(row=0, column=0, sticky='e')
    ch_mode_menu = ttk.OptionMenu(api_frame, ch_mode_var, "Custom", "Gus's API", "Blank", command=on_ch_mode_change)
    ch_mode_menu.grid(row=0, column=1, sticky='ew')

    ttk.Label(api_frame, text="Companies House API Key:").grid(row=1, column=0, sticky='e')
    api_key_entry = ttk.Entry(api_frame)
    api_key_entry.grid(row=1, column=1, sticky='ew')
    ch_link_label = tk.Label(api_frame, text="Get Companies House API Key", fg="blue", cursor="hand2", bg="#f0f0f0")
    ch_link_label.grid(row=1, column=2)
    ch_link_label.bind("<Button-1>", lambda e: webbrowser.open_new("https://developer.company-information.service.gov.uk/get-started"))

    use_cache_check = ttk.Checkbutton(api_frame, text="Use Cache", variable=use_cache_var)
    use_cache_check.grid(row=2, column=1, sticky='w')
    verbose_check_button = ttk.Checkbutton(api_frame, text="Verbose log", variable=verbose_var)
    verbose_check_button.grid(row=2, column=2, sticky='w')

    output_frame = tk.Frame(root, bg="#f0f0f0")
    output_frame.grid(row=2, column=0, sticky='ew', pady=5)
    output_frame.columnconfigure(1, weight=1)

    ttk.Label(output_frame, text="Output Filename:").grid(row=0, column=0, sticky='e')
    output_entry = ttk.Entry(output_frame, textvariable=output_var, width=50)
    output_entry.grid(row=0, column=1, sticky='ew')
    output_browse_button = ttk.Button(output_frame, text="Browse", command=browse_output)
    output_browse_button.grid(row=0, column=2)

    profiles_frame = tk.Frame(root, bg="#f0f0f0")
    profiles_frame.grid(row=3, column=0, sticky='ew', pady=5)
    profiles_frame.columnconfigure(1, weight=1)

    ttk.Label(profiles_frame, text="Load Profile:").grid(row=0, column=0, sticky='e')
    profile_menu = ttk.OptionMenu(profiles_frame, profile_var, "Default")
    profile_menu.grid(row=0, column=1, sticky='ew')
    save_profile_button = ttk.Button(profiles_frame, text="Save Profile", command=save_profile)
    save_profile_button.grid(row=0, column=2)
    delete_profile_button = ttk.Button(profiles_frame, text="Delete Profile", command=delete_profile)
    delete_profile_button.grid(row=0, column=3)

    columns_frame = tk.Frame(root, bg="#f0f0f0")
    columns_frame.grid(row=4, column=0, sticky='nsew', pady=5)
    columns_frame.columnconfigure(0, weight=1)
    columns_frame.columnconfigure(1, weight=1)
    columns_frame.rowconfigure(1, weight=1)

    ttk.Label(columns_frame, text="Output Columns:").grid(row=0, column=0, columnspan=3, sticky='ew')
    available_listbox = Listbox(columns_frame, selectmode=tk.SINGLE, height=10, bg="white")
    available_listbox.grid(row=1, column=0, sticky='nsew')
    selected_listbox = Listbox(columns_frame, selectmode=tk.SINGLE, height=10, bg="white")
    selected_listbox.grid(row=1, column=1, sticky='nsew')
    buttons_frame = tk.Frame(columns_frame, bg="#f0f0f0")
    buttons_frame.grid(row=1, column=2, sticky='ns')
    add_column_button = ttk.Button(buttons_frame, text="Add", command=add_column)
    add_column_button.pack(fill='x')
    remove_column_button = ttk.Button(buttons_frame, text="Remove", command=remove_column)
    remove_column_button.pack(fill='x')
    up_column_button = ttk.Button(buttons_frame, text="Up", command=move_up)
    up_column_button.pack(fill='x')
    down_column_button = ttk.Button(buttons_frame, text="Down", command=move_down)
    down_column_button.pack(fill='x')
    select_all_button = ttk.Button(buttons_frame, text="Select All", command=select_all_columns)
    select_all_button.pack(fill='x')

    graphs_frame = tk.Frame(root, bg="#f0f0f0")
    graphs_frame.grid(row=5, column=0, sticky='ew', pady=5)
    graphs_frame.columnconfigure(1, weight=1)

    ttk.Label(graphs_frame, text="Summary Graphs:").grid(row=0, column=0, sticky='e')
    insolvent_check = ttk.Checkbutton(graphs_frame, text="Insolvent vs Active", variable=insolvent_var)
    insolvent_check.grid(row=0, column=1, sticky='w')

    controls_frame = tk.Frame(root, bg="#f0f0f0")
    controls_frame.grid(row=6, column=0, sticky='ew', pady=5)
    controls_frame.columnconfigure(4, weight=1)

    start_button = ttk.Button(controls_frame, text="Start", command=start_processing, style='TButton')
    start_button.grid(row=0, column=0)

    pause_button = ttk.Button(controls_frame, text="Pause")
    pause_button.grid(row=0, column=1)

    stop_button = ttk.Button(controls_frame, text="Stop", style='TRed.TButton')
    stop_button.grid(row=0, column=2)

    toggle_log_button = ttk.Button(controls_frame, text="Show/Hide Log", command=toggle_log)
    toggle_log_button.grid(row=0, column=3)

    detach_button = ttk.Button(controls_frame, text="Detach Log", command=toggle_detach)
    detach_button.grid(row=0, column=4, sticky='w')

    # Status bar frame at the bottom: visually separated, holds progress and status text.
    status_frame = tk.Frame(root, bg="#e0e0e0", relief="sunken", bd=1)
    status_frame.grid(row=8, column=0, sticky='ew', pady=(2, 0))
    status_frame.columnconfigure(0, weight=1)

    # Make the initial log widget shorter (fewer lines) so it starts at about half height.
    embedded_log_text = scrolledtext.ScrolledText(root, wrap='none', bg="white", fg="black", height=8)
    embedded_log_text.grid(row=7, column=0, sticky='nsew', pady=(2, 0))
    embedded_log_text.tag_configure("green", foreground="green")
    embedded_log_text.tag_configure("red", foreground="red")
    embedded_log_text.tag_configure("yellow", foreground="orange")
    init_log_context_menu(embedded_log_text)

    active_log_text = embedded_log_text
    is_detached = False
    log_window = None
    detached_log_text = None

    load_profile_menu()
    profile_var.trace_add("write", load_profile)
    load_profile()

    # Restore basic geometry from settings if available.
    geom = settings.get("window_geometry")
    if isinstance(geom, str):
        try:
            root.geometry(geom)
        except Exception:
            pass

    # If a spreadsheet was remembered from last session, load its metadata now.
    last_sheet = settings.get("spreadsheet")
    if last_sheet and os.path.exists(last_sheet):
        load_spreadsheet(last_sheet)

    # Ensure settings are persisted on close.
    def on_close():
        persist_runtime_settings()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

    # Attach tooltips (concise explanations) to key controls.
    ToolTip(spreadsheet_entry, "Path to the Excel spreadsheet to process.")
    ToolTip(browse_file_button, "Browse for an Excel spreadsheet.")
    ToolTip(start_row_entry, "First row number to process (1 = first data row).")
    ToolTip(end_row_entry, "Last row number to process (inclusive).")
    ToolTip(name_column_menu, "Column that contains the company or debtor names.")
    ToolTip(avg_calls, "Estimated API calls per row (for ETA only).")
    ToolTip(api_key_entry, "Your Companies House API key (required to call the API).")
    ToolTip(use_cache_check, "Reuse results during this session to save API calls.")
    ToolTip(verbose_check_button, "Show extra match details in the log (for debugging).")
    ToolTip(output_entry, "Path where the enriched Excel results will be saved.")
    ToolTip(output_browse_button, "Choose where to save the result Excel file.")
    ToolTip(profile_menu, "Choose a saved output column layout.")
    ToolTip(save_profile_button, "Save the current column selection as a profile.")
    ToolTip(delete_profile_button, "Delete the selected columns profile.")
    ToolTip(available_listbox, "All available columns from the spreadsheet and lookup.")
    ToolTip(selected_listbox, "Columns that will appear in the output, in this order.")
    ToolTip(add_column_button, "Add the selected column to the output list.")
    ToolTip(remove_column_button, "Remove the selected column from the output list.")
    ToolTip(up_column_button, "Move the selected output column up.")
    ToolTip(down_column_button, "Move the selected output column down.")
    ToolTip(select_all_button, "Include all available columns in the output.")
    ToolTip(insolvent_check, "Add summary charts for insolvent vs active accounts.")
    ToolTip(start_button, "Start processing the selected rows.")
    ToolTip(pause_button, "Pause or resume processing.")
    ToolTip(stop_button, "Stop processing after the current step finishes.")
    ToolTip(toggle_log_button, "Show or hide the log panel.")
    ToolTip(detach_button, "Open the log in a separate window.")
    ToolTip(embedded_log_text, "Processing log. Right-click for copy/select/clear options.")


def process_file_cli(file_path, start_row, end_row, api_key, output_file, name_column, selected_columns=None, graph_selections=None, use_cache=True):
    """
    Non-GUI wrapper around process_file for command-line usage.
    """
    if selected_columns is None:
        selected_columns = []
    if graph_selections is None:
        graph_selections = ['Insolvent vs Active']

    if use_cache:
        global profile_cache, officers_cache, insolvency_cache, filing_cache, disqualified_cache, search_cache
        profile_cache = load_cache('profile_cache.pickle')
        officers_cache = load_cache('officers_cache.pickle')
        insolvency_cache = load_cache('insolvency_cache.pickle')
        filing_cache = load_cache('filing_cache.pickle')
        disqualified_cache = load_cache('disqualified_cache.pickle')
        search_cache = load_cache('search_cache.pickle')
    else:
        profile_cache.clear()
        officers_cache.clear()
        insolvency_cache.clear()
        filing_cache.clear()
        disqualified_cache.clear()
        search_cache.clear()

    try:
        df_input = pd.read_excel(file_path, sheet_name=0)
    except Exception as e:
        print(f"[ERROR] Failed to read file: {e}")
        return

    if name_column not in df_input.columns:
        print(f"[ERROR] No '{name_column}' column found.")
        return

    row_start = max(0, start_row - 1)
    row_end = min(len(df_input), end_row)
    row_count = row_end - row_start
    if row_count <= 0:
        print("[WARN] No rows to process for the given range.")
        return

    print(f"[INFO] Processing rows {start_row} to {row_end} ({row_count} total)")

    total_start_time = time.time()
    times = []
    results = []

    def log_func(text, tag=None):
        # Strip colour tags in CLI, just print the message.
        sys.stdout.write(text)
        sys.stdout.flush()

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {}
        for rel_index, abs_index in enumerate(range(row_start, row_end), 1):
            row = df_input.iloc[abs_index]
            debtor_name = row[name_column]
            abs_row_num = abs_index + 2  # account for header row
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_func(f"{current_time} - Submitted row {abs_row_num} ({rel_index}/{row_count}): {debtor_name}\n")
            future = executor.submit(process_row, row, api_key, log_func, name_column, threading.Event(), use_cache)
            futures[future] = (rel_index, abs_row_num, debtor_name)

        processed = 0
        for future in as_completed(futures):
            rel_index, abs_row_num, debtor_name = futures[future]
            try:
                results_list, duration = future.result()
            except Exception as e:
                print(f"[ERROR] Error processing row {abs_row_num}: {e}")
                results_list = []
                duration = 0.0

            times.append(duration)
            for new_row in results_list:
                original_address_parts = [str(new_row.get(key, '')) for key in ['Address Name or Number', 'Address Line1', 'Address Line2', 'Address Line3', 'Address Line4', 'Add Post Code'] if key in new_row and not pd.isna(new_row[key])]
                new_row['Original Address'] = ', '.join(original_address_parts)
                results.append(new_row)

            accuracy = results_list[0].get('Match Accuracy', '0%') if results_list else '0%'
            matched_name = results_list[0].get('Found Company Name', '') if results_list and accuracy != '0%' else ''
            explanation = results_list[0].get('Match Explanation', '') if results_list else ''
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if accuracy != '0%':
                company_no = results_list[0].get('Company Number', '')
                status = results_list[0].get('Insolvency Status', '')
                log_func(
                    f"{current_time} - Searched: {debtor_name} - Matched: {matched_name} "
                    f"(CN: {company_no}, Status: {status}) with accuracy {accuracy} "
                    f"- Duration: {duration:.2f} seconds | {explanation}\n"
                )
            elif explanation == 'Not a company':
                log_func(f"{current_time} - {debtor_name} - Not a company - Duration: {duration:.2f} seconds\n")
            else:
                log_func(f"{current_time} - {debtor_name} - Not Found - Duration: {duration:.2f} seconds\n")

            processed += 1
            if times and processed % 10 == 0:
                avg_time = sum(times) / len(times)
                remaining = row_count - processed
                eta = avg_time * remaining
                elapsed = time.time() - total_start_time
                print(f"[PROGRESS] {processed}/{row_count} rows | Elapsed: {elapsed:.1f}s | ETA: {eta:.1f}s")

    try:
        save_results(results, output_file, name_column, selected_columns, graph_selections)
    except Exception as e:
        print(f"[ERROR] Failed to save results to {output_file}: {e}")
        return
    if use_cache:
        save_caches()
    total_duration = time.time() - total_start_time
    total, matched, insolvent = compute_summary_stats(results)
    unmatched = total - matched
    print(f"[INFO] Done! Processed {row_count} rows in {total_duration:.1f}s. Results saved to {output_file}")
    print(f"[SUMMARY] Total rows: {total} | Matched: {matched} | Insolvent/flagged: {insolvent} | Unmatched/non-companies: {unmatched}")


def parse_cli_args(argv):
    parser = argparse.ArgumentParser(description="Companies House bulk lookup tool (CLI mode)")
    parser.add_argument("--cli", action="store_true", help="Run in CLI mode instead of GUI")
    parser.add_argument("--input", "-i", required=True, help="Input Excel file path")
    parser.add_argument("--output", "-o", required=True, help="Output Excel file path")
    parser.add_argument("--name-column", "-c", required=True, help="Column header containing company names")
    parser.add_argument("--start-row", type=int, default=1, help="1-based start row (default: 1)")
    parser.add_argument("--end-row", type=int, default=10**9, help="1-based end row (default: until end)")
    parser.add_argument("--api-key", help="Companies House API key (overrides env/config)")
    parser.add_argument("--use-cache", dest="use_cache", action="store_true", help="Use cached API responses")
    parser.add_argument("--no-cache", dest="use_cache", action="store_false", help="Do not use cached API responses")
    parser.set_defaults(use_cache=True)
    parser.add_argument(
        "--columns",
        nargs="+",
        help="Optional list of output columns to include in order (default: all available columns)"
    )
    return parser.parse_args(argv)


def main_cli(argv=None):
    args = parse_cli_args(argv or sys.argv[1:])

    api_key = args.api_key or GUS_API_KEY
    if not api_key:
        print("[ERROR] No API key provided. Use --api-key or set GUS_API_KEY / config.ini.")
        sys.exit(1)

    selected_columns = args.columns if args.columns else []

    process_file_cli(
        file_path=args.input,
        start_row=args.start_row,
        end_row=args.end_row,
        api_key=api_key,
        output_file=args.output,
        name_column=args.name_column,
        selected_columns=selected_columns,
        graph_selections=['Insolvent vs Active'],
        use_cache=args.use_cache,
    )


if __name__ == "__main__":
    # If --cli is present, run in CLI mode; otherwise launch the GUI.
    if "--cli" in sys.argv:
        main_cli()
    else:
        create_gui()
        root.mainloop()
